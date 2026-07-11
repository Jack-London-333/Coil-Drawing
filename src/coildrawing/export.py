"""结果导出：Excel / CSV / 大样图 PNG / STEP。"""

from __future__ import annotations

import csv

from .engine import CoilInput, CoilResult


def input_rows(inp: CoilInput) -> list[tuple[str, str, str, str]]:
    """输入参数表 (符号, 名称, 数值, 单位)。"""
    rows: list[tuple[str, str, str, str]] = [
        ("D2", "定子铁芯内径", f"{inp.d2:g}", "mm"),
        ("LC", "铁芯轴向长度", f"{inp.lc:g}", "mm"),
        ("NS", "定子槽数", str(inp.ns), ""),
        ("2P", "极数", str(inp.poles), ""),
        ("TAW", "线圈节距", str(inp.taw), "槽"),
        ("HS", "槽深", f"{inp.hs:g}", "mm"),
        ("WS", "槽宽", f"{inp.ws:g}", "mm"),
        ("HSD", "槽楔厚度", f"{inp.hsd:g}", "mm"),
        ("WIHU", "槽楔下垫片厚度", f"{inp.wihu:g}", "mm"),
        ("WIHM", "层间垫片厚度", f"{inp.wihm:g}", "mm"),
        ("WIHB", "槽底垫片厚度", f"{inp.wihb:g}", "mm"),
        ("N", "线圈匝数", str(inp.n_turns), ""),
        ("WB1×WT1", "导线1裸线 宽×厚",
         f"{inp.wire1.b:g} × {inp.wire1.h:g}", "mm"),
        ("T01", "导线1自身绝缘(单边)", f"{inp.wire1.t0:g}", "mm"),
        ("NPD1/NCD1", "导线1并绕根数/层数",
         f"{inp.wire1.npd} / {inp.wire1.ncd}", ""),
        ("WB2×WT2", "导线2裸线 宽×厚",
         f"{inp.wire2.b:g} × {inp.wire2.h:g}", "mm"),
        ("T02", "导线2自身绝缘(单边)", f"{inp.wire2.t0:g}", "mm"),
        ("NPD2/NCD2", "导线2并绕根数/层数",
         f"{inp.wire2.npd} / {inp.wire2.ncd}", ""),
        ("T1", "槽内匝间绝缘(单边)", f"{inp.t1:g}", "mm"),
        ("T3", "端部匝间绝缘(单边)", f"{inp.t3:g}", "mm"),
        ("T2", "槽内对地绝缘(单边)", f"{inp.t2:g}", "mm"),
        ("T4", "端部对地绝缘(单边)", f"{inp.t4:g}", "mm"),
        ("CS", "槽内防晕层(单边)", f"{inp.cs:g}", "mm"),
        ("LD", "齿压板轴向长度", f"{inp.ld:g}", "mm"),
        ("LE", "直线部伸出铁芯长度", f"{inp.le:g}", "mm"),
        ("F", "鼻端抬高", f"{inp.f_nose:g}", "mm"),
        ("seita3", "鼻端中心线夹角", f"{inp.seita3:g}", "rad"),
        ("RD", "鼻端半径", f"{inp.rd_nose:g}", "mm"),
        ("RD1", "接线侧弯弧半径", f"{inp.rd1_conn:g}", "mm"),
        ("RD2", "非接线侧弯弧半径", f"{inp.rd2_nonconn:g}", "mm"),
        ("rd1", "直线部-斜边弯曲半径", f"{inp.r_bend_slot:g}", "mm"),
        ("rd2", "斜边-鼻端弯曲半径", f"{inp.r_bend_nose:g}", "mm"),
        ("Ba", "端部间隙给定值", f"{inp.ba:g}", "mm"),
        ("ysc", "引线长", f"{inp.ysc:g}", "mm"),
        ("ξ", "迭代误差设定值", f"{inp.xi:g}", ""),
    ]
    for i, layer in enumerate(inp.layers, start=1):
        rows.append((f"对地层{i}", layer.name, f"{layer.thickness:g}", "mm"))
    for i, layer in enumerate(inp.turn_layers, start=1):
        rows.append((f"匝绝缘层{i}", layer.name, f"{layer.thickness:g}", "mm"))
    rows += [
        ("rlead", "引线折弯半径", f"{inp.lead_bend_r:g}", "mm"),
        ("Lbare", "引线端头裸铜长", f"{inp.lead_bare:g}", "mm"),
        ("防晕", "槽部防晕层(3D，厚度=CS)", "启用" if inp.corona_on else "关闭", ""),
    ]
    if inp.corona_on:
        rows.append(
            ("Lcor+", "防晕层每端伸出铁芯(沿导线)", f"{inp.corona_overhang:g}", "mm"))
    rows.append(("固定件", "3D 槽内固定件",
                 "、".join(x for on, x in ((inp.draw_wedge, "槽楔"),
                                          (inp.draw_wihu, "槽楔下垫片"),
                                          (inp.draw_wihm, "层间垫片"),
                                          (inp.draw_wihb, "槽底垫片")) if on) or "无", ""))
    return rows


def export_csv(res: CoilResult, filepath: str) -> None:
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["类别", "符号", "名称", "数值", "单位"])
        for row in input_rows(res.inp):
            w.writerow(["输入", *row])
        for row in res.rows():
            w.writerow(["结果", *row])
        for warn in res.warnings:
            w.writerow(["警告", "", warn, "", ""])


def export_xlsx(res: CoilResult, filepath: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    def fill_sheet(ws, rows, title):
        ws.append([title, "", "", ""])
        ws["A1"].font = Font(bold=True, size=13)
        ws.append(["符号", "名称", "数值", "单位"])
        head_fill = PatternFill("solid", fgColor="DDEBF7")
        for c in ws[2]:
            c.font = Font(bold=True)
            c.fill = head_fill
        sec_fill = PatternFill("solid", fgColor="FFF2CC")
        for sym, name, val, unit in rows:
            ws.append([sym, name, val, unit])
            if sym == "—":  # 分节行
                for c in ws[ws.max_row]:
                    c.fill = sec_fill
                    c.font = Font(bold=True)
        widths = [14, 34, 16, 8]
        for i, w_ in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w_
        for row in ws.iter_rows(min_row=2):
            row[2].alignment = Alignment(horizontal="right")

    ws1 = wb.active
    ws1.title = "输入参数"
    fill_sheet(ws1, input_rows(res.inp), "输入参数")

    ws2 = wb.create_sheet("计算结果")
    fill_sheet(ws2, res.rows(), "计算结果（CN104965948B 步骤1-6）")

    if res.warnings:
        ws3 = wb.create_sheet("警告")
        ws3.append(["警告信息"])
        ws3["A1"].font = Font(bold=True, color="CC0000")
        for warn in res.warnings:
            ws3.append([warn])
        ws3.column_dimensions["A"].width = 100

    wb.save(filepath)
